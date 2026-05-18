from __future__ import annotations

import sqlite3
import sys
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
DB_PATH = ROOT / "data" / "gumus_veteriner.db"


def clean(value) -> str:
    return "" if value is None else str(value).strip()


def to_float(value) -> float:
    if value in (None, ""):
        return 0.0
    return float(str(value).replace(",", "."))


def to_int(value) -> int:
    if value in (None, ""):
        return 0
    return int(float(str(value).replace(",", ".")))


def reset_products_table(db: sqlite3.Connection) -> None:
    db.executescript(
        """
        DROP TABLE IF EXISTS products;

        CREATE TABLE products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            category TEXT NOT NULL,
            price REAL NOT NULL,
            stock INTEGER NOT NULL,
            image_url TEXT,
            description TEXT,
            image_emoji TEXT,
            active INTEGER NOT NULL DEFAULT 1
        );
        """
    )


def import_products(xlsx_path: Path) -> int:
    workbook = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [clean(cell.value) for cell in sheet[1]]
    index = {name: i for i, name in enumerate(headers)}

    required = ["Kategori Adı", "Ürün Adı", "Trendyol Satış Fiyatı", "Stok", "Görsel 1"]
    missing = [name for name in required if name not in index]
    if missing:
        raise RuntimeError(f"Eksik kolonlar: {', '.join(missing)}")

    rows = []
    for raw in sheet.iter_rows(min_row=2, values_only=True):
        name = clean(raw[index["Ürün Adı"]])
        if not name:
            continue
        rows.append(
            (
                name,
                clean(raw[index["Kategori Adı"]]) or "Genel",
                to_float(raw[index["Trendyol Satış Fiyatı"]]),
                to_int(raw[index["Stok"]]),
                clean(raw[index["Görsel 1"]]),
                1,
            )
        )

    DB_PATH.parent.mkdir(exist_ok=True)
    with sqlite3.connect(DB_PATH) as db:
        reset_products_table(db)
        db.executemany(
            """
            INSERT INTO products (name, category, price, stock, image_url, active)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            rows,
        )
        db.commit()
    return len(rows)


if __name__ == "__main__":
    source = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(r"C:\Users\ghost\Desktop\ürün stok takip.xlsx")
    total = import_products(source)
    print(f"{total} ürün içe aktarıldı: {DB_PATH}")
